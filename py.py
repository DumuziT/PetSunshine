import openpyxl, os
from openpyxl.styles import Border, Side, PatternFill, Font

# Crea un nuevo libro de trabajo
name = str(input("Nombre del archivo: "))

Materiales = []
Mcantidad = []
MvalorUnitario = []
MvalorTotal = []

Trabajos = []
Tcantidad = []
TvalorUnitario = []
TvalorTotal = []

FilaMaterial = 0
FilaTrabajo = 0

while True:
  os.system("clear")
  print("""1) Agregar fila\n2) Eliminar fila\n3) Salir\n""")
  opcion = int(input("Opcion: "));
  if opcion == 1:
    print("MATERIALES")
    var_Materiales = str(input("Materiales: "))
    var_Mcantidad = int(input("Cantidad: "))
    var_MvalorUnitario = int(input("Valor Unitario: "))
    var_MvalorTotal = str(var_Mcantidad*var_MvalorUnitario)
    print("MANO DE OBRA")
    var_Trabajos = str(input("Trabajos: "))
    var_Tcantidad = int(input("Cantidad: "))
    var_TvalorUnitario = int(input("Valor Unitario: "))
    var_TvalorTotal = str(var_Tcantidad*var_TvalorUnitario)

    Materiales.append(var_Materiales)
    Mcantidad.append(var_Mcantidad)
    MvalorUnitario.append(var_MvalorUnitario)
    MvalorTotal.append(var_MvalorTotal)
    Trabajos.append(var_Trabajos)
    Tcantidad.append(var_Tcantidad)
    TvalorUnitario.append(var_TvalorUnitario)
    TvalorTotal.append(var_TvalorTotal)
    FilaMaterial = FilaMaterial+1
    FilaTrabajo = FilaTrabajo+1

  elif opcion == 2:
    print("fila|materiales|cantidad|valorUnitario|valorTotal")
    for i in range(0, FilaMaterial):
      print("{}|{}|{}|{}|{}".format(i+1, Materiales[i], Mcantidad[i], MvalorUnitario[i], MvalorTotal[i]))
    print("fila| trabajos |cantidad|valorUnitario|valorTotal")
    for j in range(0, FilaTrabajo):
      print("{}|{}|{}|{}|{}".format(j+1, Trabajos[j], Tcantidad[j], TvalorUnitario[j], TvalorTotal[j]))

    print("""\n1) Material\n2) Trabajos\nx) Cancelar""")
    Eliminar = int(input("Que quiere eliminar: "))
    eliminar = int(input("Fila que quiera eliminar: "))
    if Eliminar == 1:
      Materiales.pop(eliminar-1)
      Mcantidad.pop(eliminar-1)
      MvalorUnitario.pop(eliminar-1)
      MvalorTotal.pop(eliminar-1)
      FilaMaterial = FilaMaterial-1
    elif Eliminar == 2:
      Trabajos.pop(eliminar-1)
      Tcantidad.pop(eliminar-1)
      TvalorUnitario.pop(eliminar-1)
      TvalorTotal.pop(eliminar-1)
      FilaTrabajo = FilaTrabajo-1
    else:
      continue

  elif opcion == 3:
    exit()
  else:
    continue

# Nombre del archivo Excel plantilla
plantilla = 'aaa.xlsx'

# Cargar el libro de trabajo (workbook)
libro_trabajo = openpyxl.load_workbook(plantilla)

# Seleccionar la hoja de trabajo (worksheet) activa
hoja_trabajo = libro_trabajo.active

# Conseguir Color
colorFondo = hoja_trabajo['A1'].fill.start_color.rgb

Fila = int(input("Inserte fila nueva: "))
Text = str(input("Texto: "))

# Crear una nueva fila en la posición deseada
nueva_fila = hoja_trabajo.insert_rows(Fila)

# Realizar las operaciones de edición en la nueva fila, por ejemplo, establecer valores en celdas específicas
hoja_trabajo.cell(row=Fila, column=1, value=Text).font = Font(color='FFFFFF', name='Arial', size=10)

# Definir un fondo amarillo para todas las celdas en la nueva fila
for columna in range(1, hoja_trabajo.max_column + 1):
  celda = hoja_trabajo.cell(row=Fila, column=columna)
  celda.fill = PatternFill(start_color=colorFondo, end_color=colorFondo, fill_type='solid')
  celda.border = Border(left=Side(style='thin', color='FFFFFF'), right=Side(style='thin', color='FFFFFF'), top=Side(style='thin', color='FFFFFF'), bottom=Side(style='thin', color='FFFFFF'))

# Guardar el libro de trabajo modificado con un nuevo nombre
libro_trabajo.save(name+".xlsx")

# Cerrar el libro de trabajo
libro_trabajo.close()
